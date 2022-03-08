import pandas as pd
import os
import re
from openpyxl import load_workbook

def main(fname)

  text = ""
  data_list = []
  final_list = []

  def append_df_to_excel(filename, df, sheet_name='New', startrow=None,
                         truncate_sheet=False,
                         **to_excel_kwargs):
      if not os.path.isfile(filename):
          df.to_excel(
              filename,
              sheet_name=sheet_name,
              startrow=startrow if startrow is not None else 0,
              **to_excel_kwargs)
          return

      if 'engine' in to_excel_kwargs:
          to_excel_kwargs.pop('engine')

      writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')
      writer.book = load_workbook(filename)

      if startrow is None and sheet_name in writer.book.sheetnames:
          startrow = writer.book[sheet_name].max_row

      if truncate_sheet and sheet_name in writer.book.sheetnames:
          idx = writer.book.sheetnames.index(sheet_name)
          writer.book.remove(writer.book.worksheets[idx])
          writer.book.create_sheet(sheet_name, idx)

      writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

      if startrow is None:
          startrow = 0

      df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
      writer.save()


  with open(fname, "r", encoding="windows-1253") as text_file:
      text = text_file.readlines()

  # remove top and bottom lines
  for i in text:
      if len(i) > 120:
          if "-------" not in i:
              if "VESSEL" not in i:
                  b = re.split("GOLDEN BRIDGE ", i, maxsplit=2)
                  data_list.append(b)

  with pd.ExcelWriter("cashier.xlsx", "xlsxwriter") as writer:
      pd.DataFrame(data_list, columns=["Date", "Data"]).to_excel(writer, index=None)

  df = pd.read_excel("cashier.xlsx")

  # remove unwanted 0, prefixes and whitespace
  df[['Data', 'FNumber']] = df['Data'].str.split('00000', expand=True)
  df['Data'] = df['Data'].str[:10]
  df['Data'] = df['Data'].str.strip(" ")
  df = df.join(df['FNumber'].str.split(' +', expand=True).add_prefix('dat'))
  del df['FNumber']
  del df['dat5']

  # change column names according to needs
  df['dat1'] = df['dat1'].str.replace(",",".").astype(float)
  df['dat2'] = df['dat2'].str.replace(",",".").astype(float)
  df['dat3'] = df['dat3'].str.replace(",",".").astype(float)
  df['dat4'] = df['dat4'].str.replace(",",".").astype(float)
  df['dat0'] = df['dat0'].astype(int)
  df['Date'] = df['Date'].str[5:]
  df['Date'] = df['Date'].str.strip()
  data = df.drop_duplicates(subset=["dat0"], keep=False)


  with pd.ExcelWriter("cashier.xlsx", "xlsxwriter") as writer:
      pd.DataFrame(data).to_excel(writer, index=None)

  # group by company and user - PICK YOUR COLUMNS
  x = df.groupby(['dat1','dat2','dat3'])['Date'].count()
  append_df_to_excel('cashier.xlsx', x, sheet_name="Totals")

  clean_data = pd.read_excel("cashier.xlsx")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Filename')
    parser.add_argument('-f', type=str, required=True, help="Enter the filename of the cashier report")
    args = parser.parse_args()
    main(args.fname)
