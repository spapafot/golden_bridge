import pandas as pd
import pdfplumber
import sys
import re
import os
from openpyxl import load_workbook

listed = []
final = []
text = ""
destination_port = ""

def main(fname):
    
    fname_ = fname[:-4]
    if "IGBR" in fname:
        destination_port = "BRI"
    elif "BRIG" in fname:
        destination_port = "IGO"
    else:
        print("FILENAME MUST CONTAIN 'IGBR' or 'BRIG'")
        sys.exit()

    def append_df_to_excel(filename, df, sheet_name='New', startrow=None,
                           truncate_sheet=False,
                           **to_excel_kwargs):
        if not os.path.isfile(filename):
            df.to_excel(
                filename,
                sheet_name=sheet_name,
                startrow=startrow if startrow is not None else 0,
                **to_excel_kwargs)
            return

        if 'engine' in to_excel_kwargs: # remove unwanted args
            to_excel_kwargs.pop('engine')

        writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')
        writer.book = load_workbook(filename)

        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        if truncate_sheet and sheet_name in writer.book.sheetnames:
            idx = writer.book.sheetnames.index(sheet_name)
            writer.book.remove(writer.book.worksheets[idx])
            writer.book.create_sheet(sheet_name, idx)

        writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

        if startrow is None:
            startrow = 0

        df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)
        writer.save()

    with pdfplumber.open(fname) as pdf:
        for page in pdf.pages:
            text += page.extract_text()

    with open("sample.txt", "w") as text_file:
        text_file.write(text)

    with open("sample.txt", "r") as txt:
        fin = txt.readlines()

    for i in fin:
        try:
            i = i[2:]
            b = int(i[0])
            listed.append(i)

        # if it doesn't start with an integer it's not a valid record
        except ValueError: 
            continue
        except IndexError:
            continue

    # do not remove whitespace yet to check for page breaks based on length
    for i in listed:
        if len(i) > 30:
            i = i[2:]
            b = re.split(f"{destination_port} |,0 |EUR ", i, maxsplit=3)
            final.append(b)

    with pd.ExcelWriter("temp.xlsx", engine="xlsxwriter") as writer:
        pd.DataFrame(final, columns=["Ticket No", "Client", "Client_", "Agency"]).to_excel(writer, index=None)

    # create whatever column you need
    data_ = pd.read_excel("temp.xlsx")
    data_["Client_"] = data_["Client_"].str[-8:]
    data_["Client"] = data_["Client"].str[:-2]
    data_["Ticket No"] = data_["Ticket No"].str.replace(" ","")
    y = data_[data_['Agency'].str.contains("GOLDEN")]
    x = data_.groupby(["Agency","Client_"])["Agency"].count()

    with pd.ExcelWriter(f"{fname_}.xlsx", engine="xlsxwriter") as writer:
        pd.DataFrame(data_, columns=["Ticket No", "Client", "Client_", "Agency"]).to_excel(writer, sheet_name='All',index=None)

    # append the list and the totals in different sheets
    append_df_to_excel(f"{fname_}.xlsx", y, sheet_name="Golden")
    append_df_to_excel(f"{fname_}.xlsx", x, sheet_name="Totals")


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Filename')
    parser.add_argument('-f', type=str, required=True, help="Enter the filename in '1607BRIG.pdf' format")
    args = parser.parse_args()
    main(args.fname)
